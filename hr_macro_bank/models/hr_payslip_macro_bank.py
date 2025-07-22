from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class HrpayslipMacroBank(object):
    def __init__(self, data_by_bank, obj):
        self.data = data_by_bank
        self.obj = obj
        self.wb = Workbook()
        self.create_sheets()
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

    def create_sheets(self):
        bank_info = self._get_bank_info()
        for bank_code, bank_data in bank_info.items():
            bank_abbr = bank_data['abreviatura']
            if bank_abbr in self.data:
                self.wb.create_sheet(title=bank_abbr)

    def _get_bank_info(self):
        return {
            'BCONPEPL': {
                'nombre': 'BANCO BBVA PERU',
                'abreviatura': 'BBVA',
                'pais': 'Perú',
                'style': {
                    'font': Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
                    'fill': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),
                    'first_row_font': Font(name='Arial Black', bold=True, size=15, color='FFFFFF'),
                    'first_row_fill': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),
                },
                'columns': [
                    ('TIPO DE DOCUMENTO', 15),
                    ('NUMERO DE DOCUMENTO', 20),
                    ('NOMBRE DEL BENEFICIARIO', 50),
                    ('TIPO DE CUENTA', 15),
                    ('NUMERO DE CUENTA', 25),
                    ('MONTO', 20),
                ]
            },
            'BCPLPEPL': {
                'nombre': 'BANCO DE CREDITO DEL PERU',
                'abreviatura': 'BCP',
                'pais': 'Perú',
                'style': {
                    'font': Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
                    'fill': PatternFill(start_color='012593', end_color='012593', fill_type='solid'),
                    'first_row_font': Font(name='Arial Black', bold=True, size=15, color='FFFFFF'),
                    'first_row_fill': PatternFill(start_color='012593', end_color='012593', fill_type='solid'),
                },
                'columns': [
                    ('TIPO DE REGISTRO', 12),
                    ('TIPO DE CUENTA DE ABONO', 14),
                    ('CUENTA DE ABONO', 20),
                    ('TIPO DE DOCUMENTO DE IDENTIDAD', 15),
                    ('NUMERO DE DOCUMENTO DE IDENTIDAD', 16),
                    ('NOMBRE DEL TRABAJADOR', 50),
                    ('TIPO DE MONEDA DE ABONO', 15),
                    ('MONTO DE ABONO', 20),
                    ('VALIDACION IDC DEL PROVEEDOR VS CUENTA', 22)
                ]
            },
            'BSUDPEPL': {
                'nombre': 'SCOTIABANK PERU S.A.A.',
                'abreviatura': 'SCOTIABANK',
                'pais': 'Perú',
                'style': {
                    'font': Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
                    'fill': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
                    'first_row_font': Font(name='Arial Black', bold=True, size=15, color='FFFFFF'),
                    'first_row_fill': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
                },
                'columns': [
                    ('TIPO DE DOCUMENTO', 12),
                    ('DOCUMENTO DE IDENTIDAD', 12),
                    ('NOMBRE DEL EMPLEADO', 40),
                    ('FORMA DE PAGO', 15),
                    ('CUENTA SCOTIABANK', 25),
                    ('CUENTA INTERBANCARIA', 25),
                    ('REGIMEN LABORAL', 15),
                    ('CONCEPTO', 12),
                    ('IMPORTE', 13)
                ]
            }
        }

    def setup_headers(self, sheet, bank_abbr):
        bank_info = self._get_bank_info()
        bank = bank_info.get(bank_abbr)

        if not bank:
            return

        headers_info = bank['columns']
        style = bank['style']
        font = style['font']
        fill = style['fill']
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Asigna los encabezados y anchos de columna
        for col_num, (header, width) in enumerate(headers_info, start=1):
            cell = sheet.cell(row=10, column=col_num, value=header)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            column_letter = sheet.cell(row=10, column=col_num).column_letter
            sheet.column_dimensions[column_letter].width = width

        sheet.row_dimensions[10].height = 40

    def setup_first_row(self, sheet, bank_abbr):
        bank_info = self._get_bank_info()
        bank = bank_info.get(bank_abbr)

        if not bank:
            return

        style = bank['style']
        cell = sheet.cell(row=1, column=1, value=bank['nombre'])
        cell.font = style['first_row_font']
        cell.fill = style['first_row_fill']
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(bank['columns']))
        sheet.row_dimensions[1].height = 40

    def fill_data(self, sheet, data, bank_abbr):
        bank_info = self._get_bank_info()
        bank = bank_info.get(bank_abbr)
        if not bank:
            return

        headers_info = bank['columns']
        header_keys = [header for header, _ in headers_info]

        row_num = 11  # Primera fila después de los encabezados
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for line in data:
            for col_num, header in enumerate(header_keys, start=1):
                value = line.get(header, '')
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.border = border
            row_num += 1

    def save_excel(self, filename):
        self.wb.save(filename)

    def _get_bytes(self):
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_excel(self, data_by_bank):
        if not data_by_bank:
            raise ValueError("No se encontraron datos para los bancos")

        bank_info = self._get_bank_info()

        for bank_code, bank_data in bank_info.items():
            sheet_name = bank_data['abreviatura']
            if sheet_name in self.wb.sheetnames:
                sheet = self.wb[sheet_name]
                self.setup_first_row(sheet, bank_code)
                self.setup_headers(sheet, bank_code)

                sheet_data = data_by_bank.get(sheet_name, [])
                if sheet_data:
                    self.fill_data(sheet, sheet_data, bank_code)
                else:
                    print(f"No hay datos para la hoja {sheet_name}")

    def get_content(self):
        self.generate_excel(self.data)
        return self._get_bytes()
