from odoo import models, api
from odoo.tools.float_utils import float_is_zero
from odoo.addons.account_reports_oa.models.data_mov_year import DataMovYear
from odoo.addons.account_reports_oa.models.account_move_line import AccountMoveLine
from odoo.addons.account.models.account_partial_reconcile import AccountPartialReconcile
from odoo.addons.base.models.res_currency import CurrencyRate
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from typing import List


class ReportAnnualBalance(models.AbstractModel):
    _name = 'report.account_reports_oa.report_annual_balance'
    _inherit = 'report.report_xlsx.abstract'
    _description = 'Annual Balance Report'
    
    _init_row_ab = 8
    _init_row_group = 8
    _ns_annual_balance = "MOV ANUAL"
    _ns_det_renta = "DET RENTA"
    _ns_esf = "ESF"
    _ns_eri = "ERI"
    
    border_color = 'FFBFBFBF'
    number_format = '_-* #,##0.00_-;_-* -#,##0.00_-;_-* "-"??_-;_-@_-'
    
    line_font = Font(size=10)
    line_side = Side(style='thin', color=border_color)
    line_border = Border(top=line_side, bottom=line_side, left=line_side, right=line_side)
    line_normal_style = NamedStyle("Line Normal", font=line_font, border=line_border)
    line_currency_style = NamedStyle("Line Currency", font=line_font, border=line_border, number_format=number_format)
    
    totals_font = Font(size=10, bold=True)
    totals_title_style = NamedStyle("Totals Title", font=totals_font, alignment=Alignment(horizontal='right'))
    totals_currency_style = NamedStyle("Totals Currency", font=totals_font, number_format=number_format)
    
    # STATIC METHODS
    @staticmethod
    def _get_date_range(rec: DataMovYear):
        year = int(rec.year)
        month = int(rec.month_to)
        label_month = rec.months[month - 1].lower()
        last_day = rec.last_day_of_month(month, year)
        return f"Al {last_day.day} de {label_month} de {year}"
    
    # PRIVATE METHODS
    def _get_formula_by_group(self, rec: DataMovYear, group: str | List[str], negative=False):
        sign = '-' if negative else ''
        end_row = self._init_row_ab + len(rec.child_ids) - 1
        p_1 = f"'{self._ns_annual_balance}'!B{self._init_row_ab}:B{end_row}"
        p_2 = f"'{self._ns_annual_balance}'!S{self._init_row_ab}:S{end_row}"
        
        if isinstance(group, list):
            group_formulas = [f"SUMIF({p_1},{g},{p_2})" for g in group]
            return f"={sign}SUM({','.join(group_formulas)})"
        
        return f"={sign}SUMIF({p_1},{group},{p_2})"
    
    def _get_formula_income_tax(self, rec: DataMovYear):
        regimen = rec.company_id.income_regime
        
        if regimen == "general":
            return f"'{self._ns_det_renta}'!H29"
        elif regimen == "mype":
            return f"'{self._ns_det_renta}'!H28 + '{self._ns_det_renta}'!H29"
        elif regimen == "special":
            return 0  # TODO: Falta
        else:
            return 0
    
    def _get_name_cp(self, line: AccountMoveLine):
        serie = line.move_id.seriecomp_sunat
        num_cpe = line.move_id.numcomp_sunat
        
        if line.move_id.journal_id.type == 'purchase' and serie and num_cpe:
            return f"{serie}-{num_cpe}"
        
        return str(line.move_id.name)
    
    def _get_amount(self, line: AccountMoveLine, rec: DataMovYear):
        date_to = rec.last_day_of_month(int(rec.month_to), int(rec.year))
        
        domain_debit = [
            ('debit_move_id', '=', line.id),
            ('credit_move_id.date', '<=', date_to),
        ]
        domain_credit = [
            ('credit_move_id', '=', line.id),
            ('debit_move_id.date', '<=', date_to),
        ]
        
        partial_reconcile: AccountPartialReconcile = self.env['account.partial.reconcile']
        
        debit_result = partial_reconcile.read_group(domain_debit, ['amount:sum'], [])
        saldo_debit = debit_result[0].get('amount_sum', 0.0) if debit_result else 0.0
        credit_result  = partial_reconcile.read_group(domain_credit, ['amount:sum'], [])
        saldo_credit = credit_result[0].get('amount_sum', 0.0) if credit_result else 0.0
        
        balance = saldo_debit - saldo_credit if isinstance(saldo_debit, float) and isinstance(saldo_credit, float) else 0.0
        
        return line.debit - line.credit - balance
    
    # SECONDARY METHODS
    def process_annual_balance(self, wb: Workbook, rec: DataMovYear):
        ws = self.get_sheet(wb, self._ns_annual_balance)
        
        ws["E3"] = rec.year
        ws["E4"] = str(rec.company_id.name).upper()
        ws["E5"] = rec.company_id.vat
        
        for n_row, line in enumerate(rec.child_ids, start=self._init_row_ab):
            self.update_value(ws, n_row, "B", line.chart_group, style=self.line_normal_style)
            self.update_value(ws, n_row, "C", line.chart_code, style=self.line_normal_style)
            self.update_value(ws, n_row, "D", line.chart_name, style=self.line_normal_style)
            self.update_value(ws, n_row, "E", line.init_balance, style=self.line_currency_style)
            self.update_value(ws, n_row, "F", line.month_1, style=self.line_currency_style)
            self.update_value(ws, n_row, "G", line.month_2, style=self.line_currency_style)
            self.update_value(ws, n_row, "H", line.month_3, style=self.line_currency_style)
            self.update_value(ws, n_row, "I", line.month_4, style=self.line_currency_style)
            self.update_value(ws, n_row, "J", line.month_5, style=self.line_currency_style)
            self.update_value(ws, n_row, "K", line.month_6, style=self.line_currency_style)
            self.update_value(ws, n_row, "L", line.month_7, style=self.line_currency_style)
            self.update_value(ws, n_row, "M", line.month_8, style=self.line_currency_style)
            self.update_value(ws, n_row, "N", line.month_9, style=self.line_currency_style)
            self.update_value(ws, n_row, "O", line.month_10, style=self.line_currency_style)
            self.update_value(ws, n_row, "P", line.month_11, style=self.line_currency_style)
            self.update_value(ws, n_row, "Q", line.month_12, style=self.line_currency_style)
            self.update_value(ws, n_row, "R", 0, style=self.line_currency_style)
            self.update_value(ws, n_row, "S", f"=SUM(E{n_row}:R{n_row})", style=self.line_currency_style)
        
        row_totals = self._init_row_ab + len(rec.child_ids)
        self.update_value(ws, row_totals, "D", "TOTALES", style=self.totals_title_style)
        for col_letter in "EFGHIJKLMNOPQRS":
            self.update_value(
                ws=ws,
                row=row_totals,
                column_letter=col_letter,
                value=f"=SUM({col_letter}{self._init_row_ab}:{col_letter}{row_totals - 1})",
                style=self.totals_currency_style,
            )
    
    def process_esf(self, wb: Workbook, rec: DataMovYear):
        ws = self.get_sheet(wb, self._ns_esf)
        
        ws["B2"] = str(rec.company_id.name).upper()
        ws["B3"] = rec.company_id.vat or "S/N"
        ws['B5'] = self._get_date_range(rec)
        
        group_rel = [
            # (cell, group, negative)
            ("E10", "10", False),
            ("D11", "12", False),
            ("D12", "14", False),
            ("D13", "16", False),
            ("D14", "20", False),
            ("D15", "18", False),
            ("D23", "37", False),
            ("D24", "33", False),
            ("D25", "34", False),
            ("D26", "39", True),
            ("I11", "42", False),
            ("I11", "42", False),
            ("I12", "40", False),
            ("I13", "44", False),
            ("I14", "41", False),
            ("I15", "46", False),
            ("I19", "45", False),
            ("I23", "50", False),
            ("I24", "52", False),
            ("I25", "59", False),
        ]
        
        for cell, group, negative in group_rel:
            ws[cell] = self._get_formula_by_group(rec, group, negative=negative)
        
        ws["I12"] = f"{ws['I12'].value} + '{self._ns_eri}'!C27"
        ws["I14"] = f"{ws['I14'].value} + '{self._ns_det_renta}'!H23"
    
    def process_eri(self, wb: Workbook, rec: DataMovYear):
        ws = self.get_sheet(wb, self._ns_eri)
        
        ws["B2"] = str(rec.company_id.name).upper()
        ws["B3"] = rec.company_id.vat or "S/N"
        ws['B5'] = self._get_date_range(rec)
        
        group_rel = [
            # (cell, group, negative)
            ("C9", "70", False),            # VENTAS
            ("C10", "69", True),            # COSTO DE VENTAS
            ("C14", "94", True),            # GASTOS ADMINISTRATIVOS
            ("C15", "95", True),            # GASTOS DE VENTAS
            ("C19", ["73","75"], False),    # INGRESOS OTROS
            ("C20", "74", False),           # GASTOS OTROS
            ("C21", "77", False),           # INGRESOS FINANCIEROS
            ("C22", "97", True),            # GASTOS FINANCIEROS
        ]
        
        for cell, group, negative in group_rel:
            ws[cell] = self._get_formula_by_group(rec, group, negative=negative) 
        
        ws["C27"] = f"={self._get_formula_income_tax(rec)}"
    
    def process_det_renta(self, wb: Workbook, rec: DataMovYear):
        ws = self.get_sheet(wb, self._ns_det_renta)
        
        ws["B3"] = str(rec.company_id.name).upper()
        
        uit_value = rec.company_id.uit_id.amount
        
        ws["E23"] = float(rec.company_id.part_workers) / 100
        ws["G28"] = f"=IF(H25>15*{uit_value},15*{uit_value},H25)"
        ws["G29"] = f"=IF(H25>15*{uit_value},H25-G28,0)"
        
        if rec.company_id.income_regime == "general":
            ws.row_dimensions[28].hidden = True
    
    def process_annexes(self, wb: Workbook, rec: DataMovYear):
        for group_code in ("12", "13", "14", "16", "19", "41", "42", "43", "46", "47"):
            ws = self.get_sheet(wb, group_code)
            
            ws["B2"] = str(rec.company_id.name).upper()
            ws["B3"] = rec.company_id.vat
            ws["B5"] = self._get_date_range(rec)
            
            move_lines: AccountMoveLine = self.env['account.move.line'].search([
                ('company_id', '=', rec.company_id.id),
                ('parent_state', '=', 'posted'),
                ('group_id.code_prefix_start', '=', group_code),
            ])
            
            move_lines = move_lines.sorted(lambda line: (
                self._get_name_cp(line),
                line.partner_id.name if line.partner_id else line.move_id.partner_id.name,
                line.date
            ))
            
            for n_row, line in enumerate(move_lines, start=self._init_row_group):
                amount = self._get_amount(line, rec)
                
                if float_is_zero(amount, 2):
                    continue
                
                move_id = line.move_id
                document_type = move_id.l10n_latam_document_type_id.code if move_id.l10n_latam_document_type_id else ""
                partner_name = line.partner_id.name if line.partner_id else move_id.partner_id.name if move_id.partner_id else ""
                
                currency_rate = 1.0
                currency_name = line.currency_id.name
                
                if currency_name == "USD":
                    rcr: CurrencyRate = self.env["res.currency.rate"].search([
                        ("name", "=", line.date),
                        ("currency_id", "=", line.currency_id.id),
                        ("company_id", "=", line.company_id.id),
                    ])
                    currency_rate = rcr.inverse_company_rate
                
                self.update_value(ws, n_row, "B", line.move_id.invoice_date or line.date)
                self.update_value(ws, n_row, "C", document_type)
                self.update_value(ws, n_row, "D", self._get_name_cp(line))
                self.update_value(ws, n_row, "E", line.partner_id.vat or "")
                self.update_value(ws, n_row, "F", partner_name)
                self.update_value(ws, n_row, "G", currency_name)
                self.update_value(ws, n_row, "H", amount * currency_rate)
                self.update_value(ws, n_row, "I", currency_rate)
                self.update_value(ws, n_row, "J", f'=IF(G{n_row}="USD",H{n_row}/I{n_row},0)')
                self.update_value(ws, n_row, "K", line.account_id.code)
                self.update_value(ws, n_row, "L", line.account_id.name)
    
    # MAIN METHODS
    @api.model
    def create_workbook(self, docids, data=None):
        wb = self.get_workbook_by_template('account_reports_oa', 'annual_balance.xlsx')
        
        record: DataMovYear = self.get_record('data.mov.year', docids[0])
        
        self.process_annual_balance(wb, record)
        self.process_esf(wb, record)
        self.process_eri(wb, record)
        self.process_det_renta(wb, record)
        self.process_annexes(wb, record)
        
        return wb