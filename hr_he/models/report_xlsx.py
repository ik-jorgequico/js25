# -*- coding: utf-8 -*-
# ReportXlsx v0.1
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64


class ReportXlsx(object):
    # CONSTANTS
    row_header_1 = 6
    row_header_2 = 7
    default_value = "-"
    row_totals = 8
    
    headers = {
        "id": {
            "name": "ID",
            "width": None,
        },
        "code": {
            "name": "CÓDIGO",
            "width": None,
        },
        "type_doc": {
            "name": "TIPO DE DOCUMENTO",
            "width": 15.0,
        },
        "num_doc": {
            "name": "NÚMERO DE DOCUMENTO",
            "width": 15.0,
        },
        "first_last_name": {
            "name": "PRIMER APELLIDO",
            "width": 20.0,
        },
        "second_last_name": {
            "name": "SEGUNDO APELLIDO",
            "width": 20.0,
        },
        "first_name": {
            "name": "PRIMER NOMBRE",
            "width": 20.0,
        },
        "second_name": {
            "name": "SEGUNDO NOMBRE",
            "width": 20.0,
        },
        "structure_type_abbr": {
            "name": "TIPO DE RÉGIMEN",
            "width": None,
        },
        "salary": {
            "name": "REMUNERACIÓN BÁSICA",
            "width": 15.0,
        },
        "family_asig": {
            "name": "ASIGNACIÓN FAMILIAR",
            "width": 15.0,
        },
        "h_25": {
            "name": "HORAS 25%",
            "width": None,
        },
        "h_35": {
            "name": "HORAS 35%",
            "width": None,
        },
        "amount_25": {
            "name": "MONTO 25%",
            "width": None,
        },
        "amount_35": {
            "name": "MONTO 35%",
            "width": None,
        },
        "total_amount": {
            "name": "MONTO TOTAL",
            "width": 10.0,
        },
    }
    
    # STYLES
    side_thin = Side(style='thin') # 'medium'
    border_thin = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
    fill_primary = PatternFill(start_color='C5D9F1', fill_type='solid')
    fill_secondary = PatternFill(start_color='BFBFC0', fill_type='solid')
    align_center = Alignment(vertical='center', horizontal='center', wrap_text=True)
    title_font = Font(bold=True)
    header_font = Font(bold=True, size=10)
    line_font = Font(size=10)
    number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
    
    header_1_style = NamedStyle(name="Header 1", border=border_thin, fill=fill_primary, font=header_font, alignment=align_center)
    header_2_style = NamedStyle(name="Header 2", border=border_thin, fill=fill_secondary, font=header_font, alignment=align_center)
    title_1_style = NamedStyle(name="Title 1", font=title_font)
    line_style = NamedStyle(name="Line", alignment=align_center, font=line_font)
    totals_style = NamedStyle(name="Totals", border=border_thin, fill=fill_secondary, font=header_font)
    
    def __init__(self, obj) -> None:
        self.obj = obj
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def _update_cell(self, row: int, col: int, value=None, style: Optional[NamedStyle]=None, conta: bool=None, center: bool=None):
        _c = self.ws.cell(row, col, value)
        if style:
            if conta: style.number_format = self.number_format
            if center: style.alignment = self.align_center
            _c.style = style
    
    def _update_line_cell(self, row: int, line: dict, header_id: str, style: Optional[NamedStyle]=None, conta: bool=None):
        style = style or self.line_style
        col = self._get_col_by_header(header_id)
        val = line.get(header_id) or self.default_value
        self._update_cell(row, col, val, style, conta=conta)
    
    def _update_total_cell(self, id_header: str):
        col = self._get_col_by_header(id_header)
        col_letter = get_column_letter(col)
        start_row = f"{col_letter}{self.row_header_2 + 1}"
        end_row = f"{col_letter}{self.row_totals - 1}"
        self._update_cell(self.row_totals, col, f"=SUM({start_row}:{end_row})", self.totals_style, conta=True)
    
    def _add_merge_row(self, row: int, col_start: int, col_end: int, val=None, style: Optional[NamedStyle]=None):
        self.ws.merge_cells(f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}")
        
        cell = self.ws.cell(row, col_start, val)
        
        if style:
            cell.style = style
            for col in range(col_start, col_end + 1):
                self.ws.cell(row, col).border = style.border
    
    def _cell_total(self):
        init_col = get_column_letter(self.col)
        end_col = get_column_letter(self.max_col - 1)
        _range = f"{init_col}{self.row}:{end_col}{self.row}"
        return {self.max_col: f"=SUM(IF(ISEVEN(COLUMNS({_range})), {_range}, 0))"}

    def _get_col_by_header(self, header_id):
        return list(self.headers.keys()).index(header_id) + 1
    
    def _get_bytes_of_xlsx(self):
        output = BytesIO()
        output.seek(0)
        self.wb.save(output)
        output.seek(0)
        return base64.b64encode(output.read())
    
    def get_xlsx(self, data: List[Dict[str, Any]]):
        """ Retorna el contenido del Excel generado en Bytes para el almacenamiento en la base de datos.
        """
        # MAIN INFO
        self.ws.title = self.obj.label.upper().replace('/', '_')
        self._add_merge_row(2, 1, 2, "RAZÓN SOCIAL:", self.title_1_style)
        self._add_merge_row(3, 1, 2, "RUC:", self.title_1_style)
        self._add_merge_row(4, 1, 2, "PERIODO:", self.title_1_style)
        self._update_cell(2, 3, self.obj.company_id.name)
        self._update_cell(3, 3, self.obj.company_id.vat)
        self._update_cell(4, 3, self.obj.period)
        
        # PRIMARY HEADER
        self._add_merge_row(self.row_header_1, 1, 9, "DATOS GENERALES", self.header_1_style)
        self._add_merge_row(self.row_header_1, 10, 11, "INGRESOS AFECTOS Y NO AFECTOS", self.header_1_style)
        self._add_merge_row(self.row_header_1, 12, 15, "HORAS EXTRA", self.header_1_style)
        
        # SECONDARY HEADER
        self.ws.row_dimensions[self.row_header_2].height = 36
        for num_col, header in enumerate(self.headers.values(), start=1):
            col_width = header.get('width')
            if col_width:
                self.ws.column_dimensions[get_column_letter(num_col)].width = col_width
            self._update_cell(self.row_header_2, num_col, header.get('name'), self.header_2_style)
            
        # LINES
        for i_row, line in enumerate(data, start=self.row_header_2 + 1):
            self._update_line_cell(i_row, line, 'id')
            self._update_line_cell(i_row, line, 'code')
            self._update_line_cell(i_row, line, 'type_doc')
            self._update_line_cell(i_row, line, 'num_doc')
            self._update_line_cell(i_row, line, 'first_last_name')
            self._update_line_cell(i_row, line, 'second_last_name')
            self._update_line_cell(i_row, line, 'first_name')
            self._update_line_cell(i_row, line, 'second_name')
            self._update_line_cell(i_row, line, 'structure_type_abbr')
            self._update_line_cell(i_row, line, 'salary', conta=True)
            self._update_line_cell(i_row, line, 'family_asig', conta=True)
            self._update_line_cell(i_row, line, 'h_25', conta=True)
            self._update_line_cell(i_row, line, 'h_35', conta=True)
            self._update_line_cell(i_row, line, 'amount_25', conta=True)
            self._update_line_cell(i_row, line, 'amount_35', conta=True)
            self._update_line_cell(i_row, line, 'total_amount', conta=True)
            self.row_totals = i_row + 1
        
        # TOTALS
        self._update_cell(self.row_totals, 1, "TOTAL", self.totals_style, center=True)
        self._update_total_cell("salary")
        self._update_total_cell("family_asig")
        self._update_total_cell("h_25")
        self._update_total_cell("h_35")
        self._update_total_cell("amount_25")
        self._update_total_cell("amount_35")
        self._update_total_cell("total_amount")
        
        return self._get_bytes_of_xlsx()
