# -*- coding: utf-8 -*-
from odoo import models, api
from odoo.exceptions import ValidationError
from odoo.tools.misc import file_path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional, Any


class ReportExcelAbstract(models.AbstractModel):
    _name = 'report.report_xlsx.abstract'
    _description = 'Abstract XLSX Report'
    
    def update_value(self, ws: Worksheet, row: int, column_letter: str, value: Any=None, style: Optional[NamedStyle|str]=None):
        col_index = column_index_from_string(column_letter)
        cell = ws.cell(row, col_index, value)
        if style:
            cell.style = style
    
    @api.model
    def get_workbook_by_template(self, module_name: str, template_name: str, template_folder="static/templates"):
        template_path = file_path(f"{module_name}/{template_folder}/{template_name}")
        return load_workbook(template_path)
    
    @api.model
    def get_sheet(self, wb: Workbook, sheet_name=None):
        ws = wb[sheet_name] if sheet_name else wb.active
        
        if not ws:
            raise ValidationError("No se encontró la hoja en el Workbook")
        
        return ws
    
    @api.model
    def get_record(self, model_name: str, docid):
        record = self.env[model_name].browse(docid)
        
        if not record:
            raise ValidationError(f"No se encontró registro con id: {docid}")
        
        return record
    
    @api.model
    def create_workbook(self, docids, data=None):
        raise NotImplementedError("Subclasses must implement create_workbook method")
    